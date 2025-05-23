from bs4 import BeautifulSoup
import requests
import selenium.webdriver as webdriver
from selenium.webdriver.chrome.service import Service
import openpyxl
from openpyxl import Workbook, load_workbook
import time

def scrape_website(website):
    print("Launching chrome browser...")

    chrome_driver_path = "./chromedriver.exe"
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=Service(chrome_driver_path), options=options)

    try:
        driver.get(website)
        print("Page Loading...")
        html = driver.page_source
        time.sleep(10)

        return html
    
    finally:
        driver.quit()

def extract_body_content(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')

    prod_id = soup.select('#productContentWrapper > div.product-header.clearfix > div.col-md-2.product-keycode') 
    prod_id = prod_id[0].text.strip()

    prod_name = soup.select('#productContentWrapper > div.product-header.clearfix > div.col-md-10.product-title')
    prod_name = prod_name[0].text.strip()

    prod_description = soup.select('#longDescriptionWrapper > div')
    prod_description = prod_description[0].text.strip()
    

    base_url = f'https://web-s3-prod-pub.s3.ap-southeast-2.amazonaws.com/media/products/{prod_id}_'

    img_url = []
    for i in range(1, 10):
        image_url = f'{base_url}{i}.jpg'
        response = requests.head(image_url)
        if response.status_code == 200:
            img_url.append(image_url)
        else:
            break

    product_images = '\n'.join(f"- {url}" for url in img_url)

    body_content = (f"""
        product_id : {prod_id}
        product_name : {prod_name}
        product_description : {prod_description}
        product_image : {product_images}
        """
    )

    return [prod_id, prod_name, prod_description, product_images, body_content]

def save_to_excel(data):
    print("DATA RECEIVED:", data)
    Workbook = load_workbook("product_description.xlsx")
    sheet = Workbook["Sheet1"]

    i = sheet.max_row + 1
           
    sheet[f'A{i}'] = data[0]
    sheet[f'B{i}'] = data[1]
    sheet[f'C{i}'] = data[2]
    sheet[f'D{i}'] = data[3]
            

    Workbook.save("product_description.xlsx")
    Workbook.close()
    print("Data written successfully.")